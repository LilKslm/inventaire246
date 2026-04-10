"""
Generate the Inventaire246 Excel template for bulk inventory import.

Usage:
    python generate_inventory_template.py

Output:
    Inventaire246_Template.xlsx (in the same directory)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

CATEGORIES = [
    "Outils",
    "Arts&Crafts",
    "Camping",
    "Branche",
    "Groupe",
    "Cuisine",
    "Nourriture",
]

HEADERS = ["Nom", "Catégorie", "Emplacement", "Quantité Totale"]
COL_WIDTHS = [35, 18, 25, 18]

# Example rows to show the expected format
EXAMPLES = [
    ("Marteau", "Outils", "Bac Outils A", 5),
    ("Ciseaux", "Arts&Crafts", "Armoire Bricolage", 10),
    ("Tente 4 places", "Camping", "Garage - Étagère 2", 3),
    ("Drapeau de meute", "Branche", "Local Loups", 2),
    ("Jeu de société", "Groupe", "Salle commune", 4),
    ("Chaudron", "Cuisine", "Cuisine - Placard B", 6),
    ("Conserves de tomates", "Nourriture", "Réserve alimentaire", 20),
]


def generate():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaire"

    # ── Styles ──────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="D2D2D7"),
    )
    example_font = Font(name="Calibri", size=11, color="8E8E93", italic=True)

    # ── Headers ─────────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    # ── Example rows (grey italic so users know to delete them) ────────
    for row_idx, example in enumerate(EXAMPLES, start=2):
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.border = thin_border

    # ── Category dropdown (rows 2-500) ─────────────────────────────────
    cat_formula = '"' + ",".join(CATEGORIES) + '"'
    dv = DataValidation(
        type="list",
        formula1=cat_formula,
        allow_blank=True,
        showDropDown=False,  # False = show the dropdown arrow in Excel
    )
    dv.prompt = "Choisir une catégorie"
    dv.promptTitle = "Catégorie"
    dv.error = "Veuillez choisir une catégorie valide."
    dv.errorTitle = "Catégorie invalide"
    dv.showInputMessage = True
    dv.showErrorMessage = True
    dv.add("B2:B500")
    ws.add_data_validation(dv)

    # ── Freeze header row ───────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Save ────────────────────────────────────────────────────────────
    output = "Inventaire246_Template.xlsx"
    wb.save(output)
    print(f"Template genere: {output}")
    print(f"  {len(EXAMPLES)} exemples inclus (a supprimer avant l'import)")
    print(f"  Categories: {', '.join(CATEGORIES)}")
    print("  Menu deroulant sur la colonne Categorie (lignes 2-500)")


if __name__ == "__main__":
    generate()
